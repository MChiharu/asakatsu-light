import os
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

from flask import Flask, request, url_for, render_template_string
from openpyxl import load_workbook

import psycopg2
import io
import csv
from flask import Response

from datetime import datetime, timedelta
from datetime import time


# =========================
# Timezone (JST)
# =========================
JST = ZoneInfo("Asia/Tokyo")

def jst_now():
    return datetime.now(JST)

def jst_today():
    return jst_now().date()


# =========================
# Quiz (Excel)
# =========================
QUIZ_XLSX_PATH = "quiz_database.xlsx"   # リポジトリ直下
QUIZ_SHEET_NAME = "quiz"               # テンプレ通り

def load_quiz_bank_from_excel(path: str = QUIZ_XLSX_PATH, sheet_name: str = QUIZ_SHEET_NAME):
    """
    Excelから問題を読み込み、内部形式に変換する。
    必須列: id, question, choice1, choice2, choice3, choice4, answer
    任意列: category, explanation
    answer は 1〜4（人間に優しい）を想定し、内部では 0〜3 に変換する。
    """
    wb = load_workbook(path, data_only=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {path}. Found: {wb.sheetnames}")

    ws = wb[sheet_name]

    # 1行目: ヘッダ
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(v).strip() if v is not None else "" for v in header_row]
    col = {h: i for i, h in enumerate(headers)}

    required = ["id", "question", "choice1", "choice2", "choice3", "choice4", "answer"]
    missing = [h for h in required if h not in col]
    if missing:
        raise ValueError(f"Missing required columns in Excel header: {missing}. Header={headers}")

    quiz_bank = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        q = row[col["question"]] if col["question"] < len(row) else None
        if q is None or str(q).strip() == "":
            continue

        choices = []
        for key in ["choice1", "choice2", "choice3", "choice4"]:
            v = row[col[key]] if col[key] < len(row) else ""
            choices.append("" if v is None else str(v))

        ans_raw = row[col["answer"]] if col["answer"] < len(row) else None
        try:
            ans = int(str(ans_raw).strip())
        except Exception:
            continue

        if not (1 <= ans <= 4):
            continue

        cat = ""
        if "category" in col and col["category"] < len(row) and row[col["category"]] is not None:
            cat = str(row[col["category"]]).strip()

        exp = ""
        if "explanation" in col and col["explanation"] < len(row) and row[col["explanation"]] is not None:
            exp = str(row[col["explanation"]]).strip()

        quiz_bank.append({
            "question": str(q).strip(),
            "choices": choices,
            "answer_index": ans - 1,  # 0〜3
            "category": cat,
            "explanation": exp,
        })

    if not quiz_bank:
        raise ValueError("No valid quizzes loaded from Excel (all rows invalid or empty).")

    return quiz_bank


def get_today_quiz(quiz_bank):
    today = jst_today()
    key = today.year * 10000 + today.month * 100 + today.day
    idx = key % len(quiz_bank)
    return quiz_bank[idx]


# =========================
# Database (PostgreSQL via Render)
# =========================
def get_db_conn():
    url = os.environ.get("DATABASE_URL")
    if not url:
        raise RuntimeError("DATABASE_URL is not set (Render Environment Variables)")

    # Render等で必要になることがあるのでSSL要求を付与
    if "sslmode=" not in url:
        joiner = "&" if "?" in url else "?"
        url = url + f"{joiner}sslmode=require"

    return psycopg2.connect(url)


def init_db():
    conn = get_db_conn()
    cur = conn.cursor()

    # 起床ログ（既存）
    cur.execute("""
        CREATE TABLE IF NOT EXISTS wakeups (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            ts TEXT NOT NULL,
            day TEXT NOT NULL
        );
    """)

    # 称号マスタ
    cur.execute("""
        CREATE TABLE IF NOT EXISTS titles (
            id SERIAL PRIMARY KEY,
            code TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL,
            description TEXT NOT NULL,
            is_hidden BOOLEAN NOT NULL DEFAULT FALSE
        );
    """)

    # ユーザー称号（獲得履歴）
    cur.execute("""
        CREATE TABLE IF NOT EXISTS user_titles (
            id SERIAL PRIMARY KEY,
            user_name TEXT NOT NULL,
            title_code TEXT NOT NULL,
            acquired_day TEXT NOT NULL,
            UNIQUE(user_name, title_code)
        );
    """)

    conn.commit()
    cur.close()
    conn.close()

def seed_titles():
    titles = [
        # 連続ログイン
        ("streak_3", "3日坊主卒業", "3日連続でログインした", False),
        ("streak_7", "習慣化マスター", "7日連続でログインした", False),
        ("streak_14", "朝活職人", "14日連続でログインした", False),

        # 規則正しい生活
        ("regular_3", "規則正しい生活", "前日の起床時刻±30分以内を3日連続で達成した", False),

        # 隠し称号（今は登録だけ。判定は後で）
        ("noon_3", "昼夜逆転", "12:00以降の起床を3日以上達成した", True),
        ("earlyking_3", "早起き王", "最速起床を3日連続で達成した", True),
        ("no_sleep_3", "もしかして寝てない？", "04:00以前の起床を3日以上達成した", True),
    ]

    conn = get_db_conn()
    cur = conn.cursor()

    # 既に同じcodeがあれば何もしない（upsert）
    for code, name, desc, hidden in titles:
        cur.execute("""
            INSERT INTO titles (code, name, description, is_hidden)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (code) DO UPDATE
            SET name = EXCLUDED.name,
                description = EXCLUDED.description,
                is_hidden = EXCLUDED.is_hidden;
        """, (code, name, desc, hidden))

    conn.commit()
    cur.close()
    conn.close()

def get_user_login_days(user_name: str, limit: int = 60):
    """
    指定ユーザーのログイン日（day）を新しい順で返す。
    同一日の複数ログインは1日として扱う（DISTINCT）。
    """
    conn = get_db_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT DISTINCT day
        FROM wakeups
        WHERE name = %s
        ORDER BY day DESC
        LIMIT %s
    """, (user_name, limit))
    days = [r[0] for r in cur.fetchall()]
    cur.close()
    conn.close()
    return days


def calc_streak_days(days_desc: list[str]) -> int:
    """
    days_desc: ["2026-01-15", "2026-01-14", ...] のような降順
    連続日数を計算して返す（今日から途切れるまで）
    """
    if not days_desc:
        return 0

    streak = 1
    prev = datetime.strptime(days_desc[0], "%Y-%m-%d").date()
    for d in days_desc[1:]:
        cur = datetime.strptime(d, "%Y-%m-%d").date()
        if prev - cur == timedelta(days=1):
            streak += 1
            prev = cur
        else:
            break
    return streak


def grant_title_if_not_owned(user_name: str, title_code: str, acquired_day: str):
    """
    既に持っていたら何もしない。持っていなければ付与する。
    """
    conn = get_db_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO user_titles (user_name, title_code, acquired_day)
        VALUES (%s, %s, %s)
        ON CONFLICT (user_name, title_code) DO NOTHING
    """, (user_name, title_code, acquired_day))
    conn.commit()
    cur.close()
    conn.close()


def evaluate_and_grant_streak_titles(user_name: str, today_str: str):
    """
    連続ログイン称号（3/7/14）を判定して付与する。
    """
    days_desc = get_user_login_days(user_name, limit=60)
    streak = calc_streak_days(days_desc)

    if streak >= 3:
        grant_title_if_not_owned(user_name, "streak_3", today_str)
    if streak >= 7:
        grant_title_if_not_owned(user_name, "streak_7", today_str)
    if streak >= 14:
        grant_title_if_not_owned(user_name, "streak_14", today_str)

    return streak


def fetch_titles_with_holders():
    """
    称号一覧（titles）と保持者一覧（user_titles）を結合して返す。
    隠し称号は、保持者がいる場合のみ表示する。
    """
    conn = get_db_conn()
    cur = conn.cursor()

    # titles と user_titles を左結合して保持者をまとめる
    cur.execute("""
        SELECT
            t.code, t.name, t.description, t.is_hidden,
            ut.user_name
        FROM titles t
        LEFT JOIN user_titles ut
          ON t.code = ut.title_code
        ORDER BY t.id ASC, ut.user_name ASC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    # 整形
    titles = {}
    for code, name, desc, is_hidden, user_name in rows:
        if code not in titles:
            titles[code] = {
                "code": code,
                "name": name,
                "description": desc,
                "is_hidden": bool(is_hidden),
                "holders": []
            }
        if user_name:
            titles[code]["holders"].append(user_name)

    # 隠し称号は保持者がいないなら表示しない
    result = []
    for t in titles.values():
        if t["is_hidden"] and len(t["holders"]) == 0:
            continue
        result.append(t)

    return result


def fetch_user_titles(user_name: str):
    """
    特定ユーザーが持っている称号を返す（称号マスタ付き）
    """
    conn = get_db_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT t.code, t.name, t.description, t.is_hidden, ut.acquired_day
        FROM user_titles ut
        JOIN titles t
          ON ut.title_code = t.code
        WHERE ut.user_name = %s
        ORDER BY ut.acquired_day DESC, t.id ASC
    """, (user_name,))
    rows = cur.fetchall()
    cur.close()
    conn.close()

    return [
        {
            "code": r[0],
            "name": r[1],
            "description": r[2],
            "is_hidden": bool(r[3]),
            "acquired_day": r[4],
        }
        for r in rows
    ]

def _parse_time(ts_str: str):
    # "HH:MM:SS" 想定。 "HH:MM" しか無い場合も救う
    parts = ts_str.split(":")
    if len(parts) == 2:
        h, m = int(parts[0]), int(parts[1])
        s = 0
    else:
        h, m, s = int(parts[0]), int(parts[1]), int(parts[2])
    return time(h, m, s)


def get_user_wakeups(user_name: str, limit: int = 60):
    """
    ユーザーの起床ログを新しい順で返す（同一日複数回は最初の1件だけにする）
    返り値: [{"day": "...", "ts": "..."} ...] (降順)
    """
    conn = get_db_conn()
    cur = conn.cursor()
    # 同一日の中で最小tsを採用（＝一番早いログインをその日の起床とみなす）
    cur.execute("""
        SELECT day, MIN(ts) as ts
        FROM wakeups
        WHERE name = %s
        GROUP BY day
        ORDER BY day DESC
        LIMIT %s
    """, (user_name, limit))
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [{"day": r[0], "ts": r[1]} for r in rows]


def is_consecutive_days(days_desc: list[str], need: int) -> bool:
    """
    days_desc は降順。先頭から need 日が連続しているか。
    """
    if len(days_desc) < need:
        return False
    prev = datetime.strptime(days_desc[0], "%Y-%m-%d").date()
    for i in range(1, need):
        cur = datetime.strptime(days_desc[i], "%Y-%m-%d").date()
        if prev - cur != timedelta(days=1):
            return False
        prev = cur
    return True


def evaluate_and_grant_regular_3(user_name: str, today_str: str):
    """
    規則正しい生活：前日±30分以内の起床が3日連続
    条件を判定し、満たせば regular_3 を付与
    """
    logs = get_user_wakeups(user_name, limit=10)
    if len(logs) < 3:
        return False

    # 連続3日でなければ不成立
    days_desc = [x["day"] for x in logs]
    if not is_consecutive_days(days_desc, 3):
        return False

    # 時刻差を分で評価（前日との差が±30分以内が2回続けばOK）
    def minutes(t: time) -> int:
        return t.hour * 60 + t.minute  # 秒は丸め

    t0 = minutes(_parse_time(logs[0]["ts"]))  # 今日
    t1 = minutes(_parse_time(logs[1]["ts"]))  # 昨日
    t2 = minutes(_parse_time(logs[2]["ts"]))  # 一昨日

    ok01 = abs(t0 - t1) <= 30
    ok12 = abs(t1 - t2) <= 30

    if ok01 and ok12:
        grant_title_if_not_owned(user_name, "regular_3", today_str)
        return True
    return False


def evaluate_and_grant_noon_3(user_name: str, today_str: str):
    """
    昼夜逆転：12:00以降の起床が3日連続
    """
    logs = get_user_wakeups(user_name, limit=10)
    if len(logs) < 3:
        return False
    days_desc = [x["day"] for x in logs]
    if not is_consecutive_days(days_desc, 3):
        return False

    def is_noon(ts: str) -> bool:
        t = _parse_time(ts)
        return (t.hour >= 12)

    if all(is_noon(x["ts"]) for x in logs[:3]):
        grant_title_if_not_owned(user_name, "noon_3", today_str)
        return True
    return False


def evaluate_and_grant_no_sleep_3(user_name: str, today_str: str):
    """
    もしかして寝てない？：04:00以前の起床が3日連続
    """
    logs = get_user_wakeups(user_name, limit=10)
    if len(logs) < 3:
        return False
    days_desc = [x["day"] for x in logs]
    if not is_consecutive_days(days_desc, 3):
        return False

    def is_too_early(ts: str) -> bool:
        t = _parse_time(ts)
        # 04:00:00 以前
        return (t.hour < 4) or (t.hour == 4 and t.minute == 0 and t.second == 0)

    if all(is_too_early(x["ts"]) for x in logs[:3]):
        grant_title_if_not_owned(user_name, "no_sleep_3", today_str)
        return True
    return False


def evaluate_and_grant_earlyking_3(today_str: str):
    """
    早起き王：その日の最速起床者を3日連続で取った人に付与
    今日ログインした人だけで判定し、必要なら付与する。
    """
    # 直近3日分（today, yesterday, day-2）の最速者を取る
    today = datetime.strptime(today_str, "%Y-%m-%d").date()
    days = [(today - timedelta(days=i)).strftime("%Y-%m-%d") for i in range(3)]

    conn = get_db_conn()
    cur = conn.cursor()

    winners = []
    for d in days:
        # その日の最速(tsが最小)の name を取る（同点は名前順で1人）
        cur.execute("""
            SELECT name, MIN(ts) as ts
            FROM wakeups
            WHERE day = %s
            GROUP BY name
            ORDER BY ts ASC, name ASC
            LIMIT 1
        """, (d,))
        row = cur.fetchone()
        if not row:
            winners.append(None)
        else:
            winners.append(row[0])

    cur.close()
    conn.close()

    # 3日全部データが揃っていて、同じ人なら付与
    if all(winners) and winners[0] == winners[1] == winners[2]:
        grant_title_if_not_owned(winners[0], "earlyking_3", today_str)
        return winners[0]
    return None


def evaluate_and_grant_all_titles(user_name: str, today_str: str):
    """
    ログイン時に呼ぶ統合関数
    """
    streak = evaluate_and_grant_streak_titles(user_name, today_str)
    regular_ok = evaluate_and_grant_regular_3(user_name, today_str)
    noon_ok = evaluate_and_grant_noon_3(user_name, today_str)
    nosleep_ok = evaluate_and_grant_no_sleep_3(user_name, today_str)
    earlyking_user = evaluate_and_grant_earlyking_3(today_str)

    return {
        "streak": streak,
        "regular_ok": regular_ok,
        "noon_ok": noon_ok,
        "nosleep_ok": nosleep_ok,
        "earlyking_user": earlyking_user,
    }


# =========================
# Flask app
# =========================
app = Flask(__name__)

# 起動時に一度だけ準備
QUIZ_BANK = load_quiz_bank_from_excel()
try:
    init_db()
    seed_titles()
except Exception as e:
    print("DB init/seed failed:", repr(e))


# =========================
# HTML templates
# =========================
INDEX_HTML = """
<!doctype html>
<html>
  <head><meta charset="utf-8"><title>朝活ログイン</title></head>
  <body>
    <h1>朝活ログイン</h1>
    <p style="color:gray;">（現在の問題数：{{ quiz_count }}問）</p>
    <p>内定者限定・日替わりITクイズでログイン 🤖</p>

    {% if error %}
      <p style="color:red;">{{ error }}</p>
    {% endif %}

    <form method="post">
      <p>名前： <input type="text" name="name" required></p>

      <hr>
      <h2>今日のクイズ</h2>
      {% if quiz_category %}
        <p style="color:gray;">カテゴリ：{{ quiz_category }}</p>
      {% endif %}
      <p>{{ quiz_question }}</p>

      {% for choice in quiz_choices %}
        <label>
          <input type="radio" name="choice" value="{{ loop.index0 }}">
          {{ choice }}
        </label><br>
      {% endfor %}

      <p><button type="submit">起きた！ログインする</button></p>
    </form>

    <hr>
    <p><a href="{{ url_for('today') }}">今日のみんなの起床時間を見る</a></p>
    <p><a href="{{ url_for('history') }}">起床履歴（ヒストリー）を見る</a></p>
    <p><a href="{{ url_for('titles_page') }}">称号を見る</a></p>
  </body>
</html>
"""

RESULT_HTML = """
<!doctype html>
<html>
  <head><meta charset="utf-8"><title>判定</title></head>
  <body>
    <h1>{{ title }}</h1>
    <p>{{ message }}</p>

    {% if ok %}
      {% if explanation %}
        <hr>
        <p><b>解説</b></p>
        <p>{{ explanation }}</p>
      {% endif %}
      <p><a href="{{ url_for('today') }}">今日のみんなの起床時間へ</a></p>
      <script>
        setTimeout(() => { window.location.href = "{{ url_for('today') }}"; }, 1200);
      </script>
    {% else %}
      <p><a href="{{ url_for('index') }}">ログイン画面に戻る</a></p>
    {% endif %}
  </body>
</html>
"""

TODAY_HTML = """
<!doctype html>
<html>
  <head><meta charset="utf-8"><title>今日の起床時間</title></head>
  <body>
    <h1>今日の起床時間</h1>
    <p>日付: {{ today_str }}</p>

    {% if rows %}
      <table border="1" cellpadding="4">
        <tr><th>名前</th><th>起きた時間</th></tr>
        {% for name, ts in rows %}
          <tr><td>{{ name }}</td><td>{{ ts }}</td></tr>
        {% endfor %}
      </table>
    {% else %}
      <p>まだ誰も起きていません…？</p>
    {% endif %}

    <p><a href="{{ url_for('index') }}">ログインページに戻る</a></p>
    <p><a href="{{ url_for('history') }}">起床履歴を見る</a></p>
  </body>
</html>
"""

HISTORY_HTML = """
<!doctype html>
<html>
  <head><meta charset="utf-8"><title>起床履歴</title></head>
  <body>
    <h1>起床履歴（ヒストリー）</h1>
    <p>表示期間: {{ start_str }} 〜 {{ end_str }}</p>

    {% if rows_by_day %}
      {% for day, items in rows_by_day %}
        <h2>{{ day }}</h2>
        <ul>
          {% for name, ts in items %}
            <li>{{ ts }} - {{ name }}</li>
          {% endfor %}
        </ul>
      {% endfor %}
    {% else %}
      <p>まだ履歴がありません。</p>
    {% endif %}

    <hr>
    <p><a href="{{ url_for('index') }}">ログインページに戻る</a></p>
    <p><a href="{{ url_for('today') }}">今日の起床時間を見る</a></p>
  </body>
</html>
"""

TITLES_HTML = """
<!doctype html>
<html>
  <head>
    <meta charset="utf-8">
    <title>称号</title>
  </head>
  <body>
    <h1>🏅 称号</h1>

    <form method="get" action="{{ url_for('titles_page') }}">
      <label>名前で検索：</label>
      <input type="text" name="user" value="{{ user_query or '' }}" placeholder="例：ちはる">
      <button type="submit">検索</button>
      {% if user_query %}
        <a href="{{ url_for('titles_page') }}">（クリア）</a>
      {% endif %}
    </form>

    <hr>

    {% if user_query %}
      <h2>「{{ user_query }}」の称号</h2>
      {% if user_titles %}
        <ul>
          {% for t in user_titles %}
            <li>
              <b>{{ t.name }}</b>
              （{{ t.acquired_day }}）
              <br>
              <span style="color:gray;">{{ t.description }}</span>
            </li>
          {% endfor %}
        </ul>
      {% else %}
        <p>まだ称号がありません。</p>
      {% endif %}

      <hr>
      <h2>称号一覧（保持者）</h2>
    {% endif %}

    {% for t in titles %}
      <div style="margin-bottom: 18px;">
        <h3>🏷 {{ t.name }}</h3>
        <p style="margin-top:-8px; color:gray;">{{ t.description }}</p>

        {% if t.holders %}
          <p><b>保持者：</b>
            {{ t.holders | join(", ") }}
          </p>
        {% else %}
          <p style="color:gray;">保持者：まだいません</p>
        {% endif %}
      </div>
      <hr>
    {% endfor %}

    <p><a href="{{ url_for('index') }}">ログインページへ</a></p>
    <p><a href="{{ url_for('today') }}">今日の起床時間へ</a></p>
    <p><a href="{{ url_for('history') }}">履歴へ</a></p>
  </body>
</html>
"""


# =========================
# Routes
# =========================
@app.route("/", methods=["GET", "POST"])
def index():
    quiz = get_today_quiz(QUIZ_BANK)

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        choice_idx_str = request.form.get("choice")

        if not name:
            return render_template_string(
                INDEX_HTML,
                error="名前を入力してください。",
                quiz_question=quiz["question"],
                quiz_choices=quiz["choices"],
                quiz_category=quiz.get("category", ""),
                quiz_count=len(QUIZ_BANK),
            )

        if choice_idx_str is None:
            return render_template_string(
                INDEX_HTML,
                error="クイズの選択肢を選んでください。",
                quiz_question=quiz["question"],
                quiz_choices=quiz["choices"],
                quiz_category=quiz.get("category", ""),
                quiz_count=len(QUIZ_BANK),
            )

        try:
            choice_idx = int(choice_idx_str)
        except ValueError:
            return render_template_string(
                INDEX_HTML,
                error="選択肢が不正です。",
                quiz_question=quiz["question"],
                quiz_choices=quiz["choices"],
                quiz_category=quiz.get("category", ""),
                quiz_count=len(QUIZ_BANK),
            )

        if choice_idx != quiz["answer_index"]:
            return render_template_string(
                RESULT_HTML,
                ok=False,
                title="❌ 不正解！",
                message="もう一度考えてみよう！",
                explanation=None,
            )

        # 正解 → 起床時間を記録（JST）
        now = jst_now()
        ts_str = now.strftime("%H:%M:%S")
        day_str = now.strftime("%Y-%m-%d")

        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO wakeups (name, ts, day) VALUES (%s, %s, %s)",
            (name, ts_str, day_str),
        )
        conn.commit()
        cur.close()
        conn.close()
        
        award = evaluate_and_grant_all_titles(name, day_str)
        streak = award["streak"]

        new_msgs = []
        if award["regular_ok"]:
            new_msgs.append("🏅 規則正しい生活 を獲得！")
        if award["noon_ok"]:
            new_msgs.append("🕵 隠し称号：昼夜逆転 を獲得！")
        if award["nosleep_ok"]:
            new_msgs.append("🕵 隠し称号：もしかして寝てない？ を獲得！")
        if award["earlyking_user"] == name:
            new_msgs.append("🕵 隠し称号：早起き王 を獲得！")

        extra = ("<br>" + "<br>".join(new_msgs)) if new_msgs else ""


        return render_template_string(
            RESULT_HTML,
            ok=True,
            title="✅ ログイン成功！",
            message=f"{name} さんの起床時間（{ts_str}）を記録しました。連続ログイン：{streak}日{extra}",
            explanation=quiz.get("explanation") or None,
            )


    # GET
    return render_template_string(
        INDEX_HTML,
        error=None,
        quiz_question=quiz["question"],
        quiz_choices=quiz["choices"],
        quiz_category=quiz.get("category", ""),
        quiz_count=len(QUIZ_BANK),
    )


@app.route("/today")
def today():
    today_str = jst_today().strftime("%Y-%m-%d")

    conn = get_db_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT name, ts FROM wakeups WHERE day = %s ORDER BY ts ASC",
        (today_str,),
    )
    rows = cur.fetchall()
    cur.close()
    conn.close()

    return render_template_string(TODAY_HTML, today_str=today_str, rows=rows)


@app.route("/history")
def history():
    N_DAYS_HISTORY = 30  # 好きに変更OK（例：30日表示）

    end_date = jst_today()
    start_date = end_date - timedelta(days=N_DAYS_HISTORY - 1)

    start_str = start_date.strftime("%Y-%m-%d")
    end_str = end_date.strftime("%Y-%m-%d")

    conn = get_db_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT day, name, ts
        FROM wakeups
        WHERE day BETWEEN %s AND %s
        ORDER BY day DESC, ts ASC
    """, (start_str, end_str))
    rows = cur.fetchall()
    cur.close()
    conn.close()

    rows_by_day_dict = {}
    for day_str, name, ts in rows:
        rows_by_day_dict.setdefault(day_str, []).append((name, ts))

    rows_by_day = sorted(rows_by_day_dict.items(), key=lambda x: x[0], reverse=True)

    return render_template_string(
        HISTORY_HTML,
        rows_by_day=rows_by_day,
        start_str=start_str,
        end_str=end_str,
    )


# （確認用：必要なときだけ使って、動いたら消してOK）
@app.route("/admin/dbinfo")
def admin_dbinfo():
    conn = get_db_conn()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*), MIN(day), MAX(day) FROM wakeups")
    count, minday, maxday = cur.fetchone()
    cur.close()
    conn.close()
    return {"count": count, "min_day": minday, "max_day": maxday}


if __name__ == "__main__":
    # ローカル起動用。Renderではgunicornが起動するのでここは使われません
    app.run(host="0.0.0.0", port=5000, debug=True)

@app.route("/download/wakeups.csv")
def download_wakeups_csv():
    # クエリパラメータ（任意）
    # 1) days=30 なら直近30日
    # 2) start=YYYY-MM-DD&end=YYYY-MM-DD ならその範囲
    days = request.args.get("days", default=None, type=int)
    start = request.args.get("start", default=None, type=str)
    end = request.args.get("end", default=None, type=str)

    # 期間の決定（JST基準）
    end_date = jst_today()
    if days:
        start_date = end_date - timedelta(days=max(1, days) - 1)
        start_str = start_date.strftime("%Y-%m-%d")
        end_str = end_date.strftime("%Y-%m-%d")
    elif start and end:
        # 形式チェックは最低限（厳密にしたければ後で追加）
        start_str, end_str = start.strip(), end.strip()
    else:
        # デフォルト：直近30日
        start_date = end_date - timedelta(days=29)
        start_str = start_date.strftime("%Y-%m-%d")
        end_str = end_date.strftime("%Y-%m-%d")

    # DBから取得
    conn = get_db_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT day, ts, name
        FROM wakeups
        WHERE day BETWEEN %s AND %s
        ORDER BY day ASC, ts ASC
    """, (start_str, end_str))
    rows = cur.fetchall()
    cur.close()
    conn.close()

    # CSV生成（メモリ上）
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["day", "ts", "name"])
    writer.writerows(rows)

    csv_text = output.getvalue()
    output.close()

    # ★ここがポイント：Excel向けにUTF-8 BOM付きで返す
    csv_bytes = csv_text.encode("utf-8-sig")  # BOM付きUTF-8

    filename = f"wakeups_{start_str}_to_{end_str}.csv"
    return Response(
        csv_bytes,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@app.route("/admin/titles")
def admin_titles():
    conn = get_db_conn()
    cur = conn.cursor()
    cur.execute("SELECT code, name, is_hidden FROM titles ORDER BY id;")
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return {"titles": rows}

@app.route("/admin/user_titles")
def admin_user_titles():
    user = request.args.get("user")
    conn = get_db_conn()
    cur = conn.cursor()
    if user:
        cur.execute("""
            SELECT user_name, title_code, acquired_day
            FROM user_titles
            WHERE user_name = %s
            ORDER BY acquired_day DESC
        """, (user,))
    else:
        cur.execute("""
            SELECT user_name, title_code, acquired_day
            FROM user_titles
            ORDER BY acquired_day DESC
            LIMIT 200
        """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return {"user_titles": rows}

@app.route("/titles")
def titles_page():
    user = request.args.get("user", default=None, type=str)
    user_query = user.strip() if user else ""

    titles = fetch_titles_with_holders()

    user_titles = None
    if user_query:
        user_titles = fetch_user_titles(user_query)

    return render_template_string(
        TITLES_HTML,
        titles=titles,
        user_query=user_query,
        user_titles=user_titles,
    )
